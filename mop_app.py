
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import ttk, messagebox
import tkinter.filedialog as fd
from mop_logic import clean_and_aggregate, build_final_table_multi
from styles import (
    DARK_BG, DARK_FG, DARK_ACCENT, DARK_BTN_BG, DARK_BTN_FG, DARK_ENTRY_BG, DARK_ENTRY_FG, DARK_HIGHLIGHT,
    DARK_BTN_PREVIEW_ACTIVE, DARK_BTN_PREVIEW_INACTIVE
)

# Функция для применения стилей ttk
def apply_styles(style):
    style.theme_use('default')
    style.configure('Treeview', background=DARK_BG, foreground=DARK_FG, fieldbackground=DARK_BG, rowheight=28)
    style.configure('Treeview.Heading', background=DARK_HIGHLIGHT, foreground=DARK_FG, font=('Segoe UI', 10, 'bold'))
    style.map('Treeview', background=[('selected', DARK_HIGHLIGHT)])

class MopApp(tk.Toplevel):
    def show_floor_table(self, floor):
        # Найти DataFrame по этажу и показать предпросмотр
        if hasattr(self, 'floors') and floor in self.floors:
            df = self.floors[floor]
            # Передаём спец. флаг для синей темы
            self.show_preview(df, is_sheet_preview='blue')
        else:
            messagebox.showerror('Ошибка', f'Нет данных для этажа: {floor}')
    def apply_multiplier(self):
        # Применить множитель к выбранному этажу и обновить предпросмотр
        floor = self.typical_floor.get()
        try:
            mult = int(self.typical_mult.get())
        except Exception:
            mult = 1
        for f, v in self.multipliers.items():
            if f == floor:
                v.set(mult)
            else:
                v.set(1)
        self.recalc()
    def delete_selected_sheets(self):
        to_delete = set(name for var, name in self.sheet_vars if var.get())
        self.sheet_list = [name for name in self.sheet_list if name not in to_delete]
        self.update_sheet_list()
        if hasattr(self, 'last_merge_file_path') and self.last_merge_file_path:
            # Для МОП: просто обновить предпросмотр (можно доработать под бизнес-логику)
            pass

    def show_preview(self, preview_df, is_sheet_preview=False):
        # Определяем стиль: если is_sheet_preview == 'blue', то делаем синий фон и белый текст
        if is_sheet_preview == 'blue':
            self.style.configure('Treeview', background='#1a2340', foreground='#fff', fieldbackground='#1a2340', rowheight=28)
            self.style.configure('Treeview.Heading', background='#223', foreground='#fff', font=('Segoe UI', 10, 'bold'))
        else:
            self.style.configure('Treeview', background='#2a2d36', foreground='#000', fieldbackground='#2a2d36', rowheight=28)
            self.style.configure('Treeview.Heading', background='#444', foreground='#000', font=('Segoe UI', 10, 'bold'))
        if not is_sheet_preview:
            self.preview_df = preview_df
        self.tree.delete(*self.tree.get_children())
        self.tree['columns'] = ()
        self.tree['show'] = 'tree headings'
        if preview_df is not None and not preview_df.empty:
            columns = list(preview_df.columns)
            self.tree['columns'] = columns
            self.tree.heading('#0', text='№')
            self.tree.column('#0', anchor='center', width=60, minwidth=50, stretch=False)
            for col in preview_df.columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, anchor='w', width=200, minwidth=120)
            for idx, (_, row) in enumerate(preview_df.iterrows(), 1):
                alt = 'alt' if idx % 2 == 0 else ''
                tags = (alt,) if alt else ()
                num_text = f'{idx} ▸'
                self.tree.insert('', 'end', text=num_text, values=list(row), tags=tags)
            self.tree.tag_configure('alt', background='#23262b')
    # self.update_active_sheet_highlight()  # убрано, чтобы не было AttributeError

    def show_tree_error(self, message):
        self.tree.delete(*self.tree.get_children())
        self.tree['columns'] = ('Ошибка',)
        self.tree.heading('Ошибка', text='Ошибка')
        self.tree.column('Ошибка', anchor='center', width=400)
        self.tree.insert('', 'end', values=(message,))
    def __init__(self, master=None):
        super().__init__(master)
        self.title('Проверка Excel по этажам')
        # Центрирование окна
        window_width = 1200
        window_height = 700
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.geometry(f'{window_width}x{window_height}+{x}+{y}')
        self.configure(bg=DARK_BG)
        self.lift()
        self.attributes('-topmost', True)
        self.focus_force()
        self.after(100, lambda: self.attributes('-topmost', False))
        # --- Стилизация Treeview (тёмная тема) ---
        self.style = ttk.Style()
        apply_styles(self.style)

        # --- Основной интерфейс ---
        self.main_frame = tk.Frame(self, bg=DARK_BG)
        self.main_frame.pack(fill='both', expand=True)

        # --- Верхняя панель с кнопками ---
        top_frame = tk.Frame(self.main_frame, bg=DARK_BG)
        top_frame.pack(side='top', fill='x')
        tk.Button(top_frame, text='Открыть Excel', command=self.open_file, bg=DARK_BTN_BG, fg=DARK_BTN_FG).pack(side='left', padx=5, pady=5)
        tk.Button(top_frame, text='Обработать папку', command=self.open_folder, bg=DARK_BTN_BG, fg=DARK_BTN_FG).pack(side='left', padx=5, pady=5)
        self.btn_export = tk.Button(top_frame, text='Экспортировать в Excel', command=self.export_to_excel, state='disabled', bg='#e8ffe8', fg='#222', relief='groove')
        self.btn_export.pack(side='left', padx=5, pady=5)
        mult_frame = tk.Frame(top_frame, bg=DARK_BG)
        mult_frame.pack(side='left', padx=10)
        tk.Label(mult_frame, text='Типовой этаж:', fg=DARK_FG, bg=DARK_BG).grid(row=0, column=0)
        self.typical_floor = tk.StringVar()
        self.typical_mult = tk.StringVar(value='1')
        self.typical_floor_cb = None  # Combobox будет создан в open_file
        tk.Label(mult_frame, text='Множитель:', fg='#fff', bg='#222').grid(row=0, column=2)
        entry = tk.Entry(mult_frame, textvariable=self.typical_mult, width=5, bg='#333', fg='#fff', insertbackground='#fff')
        entry.grid(row=0, column=3, padx=5)
        tk.Button(mult_frame, text='Умножить', command=self.apply_multiplier, bg='#333', fg='#fff').grid(row=0, column=4, padx=5)
    def open_folder(self):
        import tempfile
        import os
        import pandas as pd
        folder_path = fd.askdirectory(title='Выберите папку с Excel-файлами')
        if not folder_path:
            return
        # Только если выбрана папка — объединяем и вызываем open_file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        for fname in os.listdir(folder_path):
            if fname.endswith('.xlsx') or fname.endswith('.xls'):
                fpath = os.path.join(folder_path, fname)
                try:
                    df = pd.read_excel(fpath, header=None)
                    sheet_name = os.path.splitext(fname)[0][:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                except Exception as e:
                    print(f'Ошибка при обработке файла {fname}: {e}')
                    continue
        writer.close()
        self.open_file(path=temp_path)
    # ...existing code...


    # (Удалено создание лишней левой панели предпросмотра)
    def open_file(self, path=None):
        import traceback
        # Диалог выбора файла до любых изменений интерфейса
        if path is None:
            path = fd.askopenfilename(title='Выберите Excel-файл', filetypes=[('Excel files', '*.xlsx *.xls')])
        if not path:
            return
        # Только если выбран файл — очищаем и пересоздаём интерфейс
        if hasattr(self, 'main_frame') and self.main_frame is not None:
            for child in self.main_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
        for attr in ['center_frame', 'floor_btns_frame', 'tree_frame', 'preview_label', 'tree', 'tree_scroll', 'tree_scroll_x']:
            if hasattr(self, attr):
                setattr(self, attr, None)
        try:
            xl = pd.ExcelFile(path)
            self.floors = clean_and_aggregate(xl)
            print('DEBUG self.floors:', {k: v.shape for k, v in self.floors.items()})
            self.multipliers = {f: tk.IntVar(value=1) for f in self.floors}
            floor_nums = sorted([str(f) for f in self.floors if f not in ('0', '00', '-1')], key=lambda x: (len(x), x))

            # --- Верхняя панель с кнопками и мультипликатором ---
            top_frame = tk.Frame(self.main_frame, bg=DARK_BG)
            top_frame.pack(side='top', fill='x')
            tk.Button(top_frame, text='Открыть Excel', command=self.open_file, bg=DARK_BTN_BG, fg=DARK_BTN_FG).pack(side='left', padx=5, pady=5)
            tk.Button(top_frame, text='Обработать папку', command=self.open_folder, bg=DARK_BTN_BG, fg=DARK_BTN_FG).pack(side='left', padx=5, pady=5)
            self.btn_export = tk.Button(top_frame, text='Экспортировать в Excel', command=self.export_to_excel, state='normal', bg='#e8ffe8', fg='#222', relief='groove')
            self.btn_export.pack(side='left', padx=5, pady=5)
            mult_frame = tk.Frame(top_frame, bg=DARK_BG)
            mult_frame.pack(side='left', padx=10)
            tk.Label(mult_frame, text='Типовой этаж:', fg=DARK_FG, bg=DARK_BG).grid(row=0, column=0)
            self.typical_floor = tk.StringVar()
            self.typical_mult = tk.StringVar(value='1')
            self.typical_floor_cb = ttk.Combobox(mult_frame, values=floor_nums, textvariable=self.typical_floor, state='readonly', width=8)
            self.typical_floor_cb.grid(row=0, column=1, padx=5)
            if floor_nums:
                self.typical_floor.set(floor_nums[0])
            tk.Label(mult_frame, text='Множитель:', fg=DARK_FG, bg=DARK_BG).grid(row=0, column=2)
            entry = tk.Entry(mult_frame, textvariable=self.typical_mult, width=5, bg=DARK_BTN_BG, fg=DARK_BTN_FG, insertbackground=DARK_FG)
            entry.grid(row=0, column=3, padx=5)
            tk.Button(mult_frame, text='Умножить', command=self.apply_multiplier, bg=DARK_BTN_BG, fg=DARK_BTN_FG).grid(row=0, column=4, padx=5)

            # --- КОМПАКТНЫЕ КНОПКИ В ВЕРХНЕЙ ПАНЕЛИ ---
            self.floor_btns_frame = tk.Frame(top_frame, bg=DARK_BG)
            self.floor_btns_frame.pack(side='left', padx=10)
            col0 = 0
            btn_result = tk.Button(self.floor_btns_frame, text='Результат', width=7, height=1, font=('Segoe UI', 9), command=lambda: self.show_table(), bg='#e8ffe8', fg='#222')
            btn_result.grid(row=0, column=col0, padx=2, pady=2)
            col0 += 1
            if '0' in self.floors:
                btn_svod = tk.Button(self.floor_btns_frame, text='Свод', width=7, height=1, font=('Segoe UI', 9), command=lambda: self.show_preview(self.floors['0'], is_sheet_preview='blue'), bg=DARK_BTN_PREVIEW_ACTIVE, fg=DARK_FG)
                btn_svod.grid(row=0, column=col0, padx=2, pady=2)
                col0 += 1
            basement_floors = [f for f in self.floors if str(f) in ('00', '-1')]
            if basement_floors:
                btn_basement = tk.Button(self.floor_btns_frame, text='Подвал', width=7, height=1, font=('Segoe UI', 9), command=lambda: self.show_preview(self.floors[basement_floors[0]], is_sheet_preview='blue'), bg=DARK_BTN_PREVIEW_ACTIVE, fg=DARK_FG)
                btn_basement.grid(row=0, column=col0, padx=2, pady=2)
            max_cols = 8
            for idx, f in enumerate(floor_nums):
                row = idx // max_cols + 1
                col = idx % max_cols
                btn = tk.Button(self.floor_btns_frame, text=f, width=3, height=1, font=('Segoe UI', 9), command=lambda fl=f: self.show_preview(self.floors[fl], is_sheet_preview='blue'), bg=DARK_BTN_PREVIEW_ACTIVE, fg=DARK_FG)
                btn.grid(row=row, column=col, padx=1, pady=1)

            # --- Центр — предпросмотр ---
            self.center_frame = tk.Frame(self.main_frame, bg=DARK_BG)
            self.center_frame.pack(side='left', fill='both', expand=True)
            self.preview_label = tk.Label(self.center_frame, text='Предпросмотр: Результат', bg=DARK_BG, fg=DARK_FG)
            self.preview_label.pack(anchor='nw')
            tree_frame = tk.Frame(self.center_frame, bg=DARK_BG)
            tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
            self.tree_frame = tree_frame
            self.tree = ttk.Treeview(tree_frame, show='headings')
            self.tree_scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
            self.tree_scroll_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
            self.tree.configure(yscrollcommand=self.tree_scroll.set, xscrollcommand=self.tree_scroll_x.set)
            self.tree.grid(row=0, column=0, sticky='nsew')
            self.tree_scroll.grid(row=0, column=1, sticky='ns')
            self.tree_scroll_x.grid(row=1, column=0, sticky='ew')
            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)
            self.recalc()
            print('DEBUG final_df:', getattr(self, 'final_df', None))
            if hasattr(self, 'btn_export') and self.btn_export:
                self.btn_export.config(state='normal')
        except Exception as e:
            tb = traceback.format_exc()
            messagebox.showerror('Ошибка', f'Не удалось загрузить файл:\n{e}\n\nTraceback:\n{tb}')
            print('ERROR:', tb)
    def recalc(self):
        mults = {f: v.get() if hasattr(v, 'get') else v for f, v in self.multipliers.items()}
        self.final_df = build_final_table_multi(self.floors, mults)
        self.show_table()
    def show_table(self):
        for widget in self.tree_frame.winfo_children():
            widget.destroy()
        if self.final_df is None or self.final_df.empty:
            tk.Label(self.tree_frame, text='Нет данных для отображения.', fg='#222', bg='#fff').pack()
            return
        columns = list(self.final_df.columns)
        columns_with_idx = ['№'] + columns
        self.tree = ttk.Treeview(self.tree_frame, columns=columns_with_idx, show='headings', height=25)
        self.style = ttk.Style()
        self.style.theme_use('default')
        self.style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#444", foreground="#000")
        self.style.configure("Treeview", font=("Segoe UI", 10), rowheight=24, borderwidth=0, background="#23262b", fieldbackground="#23262b", foreground="#000")
        self.style.map("Treeview", background=[('selected', '#444')])
        for col in columns_with_idx:
            if col == '№':
                self.tree.heading(col, text=col, anchor='center')
                self.tree.column(col, width=40, anchor='center', minwidth=30, stretch=False)
            else:
                self.tree.heading(col, text=col, anchor='center')
                self.tree.column(col, width=110, anchor='center', minwidth=60, stretch=True)
        yscroll = ttk.Scrollbar(self.tree_frame, orient='vertical', command=self.tree.yview)
        xscroll = ttk.Scrollbar(self.tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.pack(side='left', fill='both', expand=True)
        yscroll.pack(side='right', fill='y')
        xscroll.pack(side='bottom', fill='x')
        for i, row in self.final_df.iterrows():
            vals = [i+1]
            for v in row:
                if isinstance(v, float):
                    vals.append(f'{v:.1f}')
                else:
                    vals.append(v)
            check = abs(row.get('Проверка', 0))
            if check > 1e-2:
                tag = 'err'
            elif check > 1e-6:
                tag = 'warn'
            else:
                tag = 'ok'
            self.tree.insert('', 'end', values=vals, tags=(tag,))
        self.tree.tag_configure('err', background='#ffe5ec')
        self.tree.tag_configure('warn', background='#fffbe5')
        self.tree.tag_configure('ok', background='#e8ffe8')
        for col in columns_with_idx:
            if col not in ('Марка', 'Наименование', '№'):
                self.tree.column(col, anchor='e')
    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        self.output_path = fd.asksaveasfilename(title='Сохранить результат как...', defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
        if not self.output_path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = 'Сводная'
        ws.append(list(self.final_df.columns))
        for i, row in self.final_df.iterrows():
            ws.append([v if not isinstance(v, float) else round(v, 1) for v in row])
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for r in range(2, ws.max_row+1):
            val = ws.cell(row=r, column=ws.max_column).value
            try:
                if abs(float(val)) > 1e-6:
                    ws.cell(row=r, column=ws.max_column).fill = red_fill
            except:
                pass
        for f in sorted(self.floors, key=lambda x: (len(str(x)), str(x))):
            df = self.floors[f]
            ws_floor = wb.create_sheet(title=f'{f}')
            ws_floor.append(list(df.columns))
            for _, row in df.iterrows():
                ws_floor.append([v if not isinstance(v, float) else round(v, 1) for v in row])
        wb.save(self.output_path)
        messagebox.showinfo('Готово', f'Результат сохранён: {self.output_path}')

if __name__ == '__main__':
    app = MopApp()
    app.mainloop()
