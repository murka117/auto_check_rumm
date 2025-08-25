
import re
import numpy as np
from datetime import datetime

def extract_name_and_qty_columns(df):
    # Ключевые слова для поиска
    import re
    norm_cols = [str(c).replace(' ', '').lower() for c in df.columns]
    mark_idx = None
    for i, c in enumerate(norm_cols):
        if 'марка' in c:
            mark_idx = i
            break
    if mark_idx is not None and mark_idx + 2 < len(df.columns):
        name_col = df.columns[mark_idx + 1]
        qty_col = df.columns[mark_idx + 2]
        print(f"DEBUG: Марка найдена, используем {name_col} как Наименование и {qty_col} как Кол-во")
    else:
        # Найти первый числовой столбец (без букв, кроме точки/запятой, допускается дата)
        qty_col = None
        name_col = None
        for i, col in enumerate(df.columns):
            # Проверяем, что большинство значений в столбце — числа или даты
            vals = df[col].dropna().astype(str)
            num_like = vals.apply(lambda v: bool(re.match(r'^\s*\d+[\d.,/\s]*$', v)))
            if num_like.sum() > len(vals) // 2:
                qty_col = col
                if i > 0:
                    name_col = df.columns[i-1]
                break
        print(f"DEBUG: Марка не найдена, используем {name_col} как Наименование и {qty_col} как Кол-во")
    name_series = df[name_col].astype(str).str.strip() if name_col else pd.Series(dtype=str)
    qty_series = df[qty_col].apply(smart_number) if qty_col else pd.Series(dtype=float)
    result = pd.DataFrame({'Наименование': name_series, 'Кол-во': qty_series})
    result = result.dropna(how='all')
    result = result[(result['Наименование'].notna()) & (result['Кол-во'].notna())]
    return result

def normalize_key(s):
    s = str(s).strip().lower()
    s = re.sub(r'[\u2013\u2014\u2212]', '-', s)
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', s)
    return s

def smart_number(val):
    if pd.isnull(val):
        return np.nan
    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)
    s = str(val).strip().replace(',', '.').replace(' ', '')
    try:
        if s.isdigit() and 35000 < int(s) < 50000:
            dt = datetime(1899, 12, 30) + pd.to_timedelta(int(s), unit='D')
            return float(dt.day)
    except Exception:
        pass
    try:
        return float(s)
    except Exception:
        pass
    # Если не число — вернуть np.nan
    return np.nan
def clean_dataframe(df):
    # Ключевые слова для поиска нужных столбцов
    # Привести все имена столбцов к нижнему регистру без пробелов для поиска
    # Если первый столбец — марка, всё равно ищем только наименование и кол-во
    return extract_name_and_qty_columns(df)
import os
import pandas as pd
from openpyxl import load_workbook


def get_sheet_names_from_file(file_path):
    wb = load_workbook(file_path, read_only=True)
    return [f'{sheet}' for sheet in wb.sheetnames]


def get_sheet_names_from_folder(folder_path):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    sheet_names = []
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        wb = load_workbook(file_path, read_only=True)
        for sheet in wb.sheetnames:
            sheet_names.append(f'{file} | {sheet}')
    return sheet_names


def merge_sheets_in_file(file_path, sheet_list):
    wb = load_workbook(file_path, read_only=True)
    preview_rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_list:
            continue
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df_clean = clean_dataframe(df)
        if not df_clean.empty:
            df_clean['source_sheet'] = sheet_name
            preview_rows.append(df_clean)
    preview_df = pd.concat(preview_rows, ignore_index=True) if preview_rows else pd.DataFrame()
    # После объединения: если одинаковое наименование встречается на разных листах — оставить только одну строку
    if not preview_df.empty:
        def filter_across_sheets(group):
            if group['source_sheet'].nunique() == 1:
                return group
            else:
                # Оставить первую с ненулевым количеством, иначе первую попавшуюся
                nonzero = group[group['Кол-во'] != 0]
                return nonzero.iloc[[0]] if not nonzero.empty else group.iloc[[0]]
        preview_df = preview_df.groupby('Наименование', group_keys=False).apply(filter_across_sheets).reset_index(drop=True)
        preview_df = preview_df.drop(columns=['source_sheet'])
    return preview_df


def merge_all_files_in_folder(folder_path, sheet_list):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    preview_rows = []
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        wb = load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet_full_name = f'{file} | {sheet_name}'
            if sheet_full_name not in sheet_list:
                continue
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df_clean = clean_dataframe(df)
            if not df_clean.empty:
                df_clean['source_sheet'] = sheet_full_name
                preview_rows.append(df_clean)
    preview_df = pd.concat(preview_rows, ignore_index=True) if preview_rows else pd.DataFrame()
    # После объединения: если одинаковое наименование встречается на разных листах — оставить только одну строку
    if not preview_df.empty:
        def filter_across_sheets(group):
            if group['source_sheet'].nunique() == 1:
                return group
            else:
                nonzero = group[group['Кол-во'] != 0]
                return nonzero.iloc[[0]] if not nonzero.empty else group.iloc[[0]]
        preview_df = preview_df.groupby('Наименование', group_keys=False).apply(filter_across_sheets).reset_index(drop=True)
        preview_df = preview_df.drop(columns=['source_sheet'])
    return preview_df


def preview_merge_file(file_path, sheet_list):
    wb = load_workbook(file_path, read_only=True)
    preview_rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_list:
            continue
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df_clean = clean_dataframe(df)
        if not df_clean.empty:
            preview_rows.append(df_clean)
    preview_df = pd.concat(preview_rows, ignore_index=True) if preview_rows else pd.DataFrame()
    return preview_df


def preview_merge_folder(folder_path, sheet_list):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    preview_rows = []
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        wb = load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet_full_name = f'{file} | {sheet_name}'
            if sheet_full_name not in sheet_list:
                continue
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df_clean = clean_dataframe(df)
            if not df_clean.empty:
                preview_rows.append(df_clean)
    preview_df = pd.concat(preview_rows, ignore_index=True) if preview_rows else pd.DataFrame()
    return preview_df
